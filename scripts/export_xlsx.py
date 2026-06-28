"""Exporta los contratos de la base de datos a un archivo XLSX trazable.

Genera tres pestanas:
  Contratos   - tabla completa con hyperlinks a la fuente oficial (SERCOP)
  Resumen     - estadisticas, top instituciones y proveedores
  Anomalias   - contratos con alertas explicadas

Uso:
    python scripts/export_xlsx.py
    python scripts/export_xlsx.py --include-demo
    python scripts/export_xlsx.py --output mi_reporte.xlsx
"""
import argparse
import os
import sys
from collections import defaultdict
from datetime import datetime
from decimal import Decimal
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.database import SessionLocal, init_db
from models.anomaly import Anomaly
from models.contract import Contract
from models.institution import Institution
from models.supplier import Supplier
from pipelines.detect_anomalies import detect_contract_anomalies
from sqlalchemy.orm import joinedload

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------------
# Paleta y estilos
# ------------------------------------------------------------------
C_DARK = "0F172A"
C_WHITE = "FFFFFF"
C_BLUE = "2563EB"
C_BLUE_LIGHT = "EFF6FF"
C_ANOMALY = "FEE2E2"       # rojo palido
C_ANOMALY_HIGH = "FCA5A5"  # rojo mas intenso para severidad alta
C_AMBER = "FEF3C7"         # amarillo palido para severidad media
C_ALT = "F8FAFC"           # gris muy suave para filas alternas
C_BORDER = "E2E8F0"
C_GRAY_HEADER = "475569"

FONT = "Arial"


def _hdr(text: str, size: int = 10, bold: bool = True, color: str = C_WHITE) -> Font:
    return Font(name=FONT, size=size, bold=bold, color=color)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def _border(color: str = C_BORDER) -> Border:
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=False)


def _wrap() -> Alignment:
    return Alignment(horizontal="left", vertical="top", wrap_text=True)


# ------------------------------------------------------------------
# Anomalias: leer de la tabla (persistidas por detect_anomalies)
# ------------------------------------------------------------------
def _load_anomalies_from_db(db) -> list[dict]:
    rows = db.query(Anomaly).all()
    return [
        {
            "contract_id": a.contract_id,
            "type": a.anomaly_type,
            "severity": a.severity,
            "reason": a.reason,
        }
        for a in rows
    ]


# ------------------------------------------------------------------
# Hoja 1: Contratos
# ------------------------------------------------------------------
HEADERS_CONTRATOS = [
    ("#", 5),
    ("OCID / ID Externo", 36),
    ("Titulo del Contrato", 48),
    ("Institucion Compradora", 30),
    ("Proveedor", 30),
    ("RUC / Tax ID", 16),
    ("Monto (USD)", 16),
    ("Tipo de Procedimiento", 26),
    ("Fecha Adjudicacion", 18),
    ("Origen de Datos", 18),
    ("Solo Demo", 10),
    ("Fuente Oficial (URL)", 20),
    ("Anomalia", 10),
    ("Detalle de Anomalia", 42),
]


def _write_contratos(
    ws,
    contracts: list[Contract],
    anomaly_ids: set[int],
    anomaly_reasons: dict[int, list[str]],
    anomaly_severity: dict[int, str],
) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    # Cabecera
    for col_idx, (label, width) in enumerate(HEADERS_CONTRATOS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = _hdr(label, size=9)
        cell.fill = _fill(C_DARK)
        cell.alignment = _center()
        cell.border = _border("334155")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28

    for row_idx, contract in enumerate(contracts, start=2):
        is_anomaly = contract.id in anomaly_ids
        severity = anomaly_severity.get(contract.id, "")

        if is_anomaly and severity == "high":
            row_fill = _fill(C_ANOMALY)
        elif is_anomaly:
            row_fill = _fill(C_AMBER)
        elif row_idx % 2 == 0:
            row_fill = _fill(C_ALT)
        else:
            row_fill = _fill(C_WHITE)

        base_font = Font(name=FONT, size=9)

        values = [
            row_idx - 1,
            contract.external_id,
            contract.title,
            contract.institution.name if contract.institution else "",
            contract.supplier.name if contract.supplier else "",
            contract.supplier.tax_id if contract.supplier else "",
            float(contract.amount) if contract.amount is not None else None,
            contract.procedure_type or "",
            contract.award_date.isoformat() if contract.award_date else "",
            contract.data_origin,
            "Si" if contract.is_demo else "No",
            "",  # URL handled separately
            "SI" if is_anomaly else "No",
            "; ".join(anomaly_reasons.get(contract.id, [])),
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = base_font
            cell.fill = row_fill
            cell.border = _border()
            cell.alignment = Alignment(
                horizontal="right" if col_idx == 7 else "left",
                vertical="center",
            )

        # Monto: formato numerico
        amount_cell = ws.cell(row=row_idx, column=7)
        if amount_cell.value is not None:
            amount_cell.number_format = '#,##0.00'

        # Anomalia: color de texto
        anomaly_flag_cell = ws.cell(row=row_idx, column=13)
        if is_anomaly:
            anomaly_flag_cell.font = Font(name=FONT, size=9, bold=True, color="B91C1C")
            anomaly_flag_cell.alignment = _center()

        # URL: hyperlink clickeable
        url_cell = ws.cell(row=row_idx, column=12)
        if contract.source_url:
            url_cell.value = "Ver en SERCOP"
            url_cell.hyperlink = contract.source_url
            url_cell.font = Font(name=FONT, size=9, color=C_BLUE, underline="single")
            url_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[row_idx].height = 16

    # Borde de tabla
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS_CONTRATOS))}1"


# ------------------------------------------------------------------
# Hoja 2: Resumen
# ------------------------------------------------------------------
def _write_resumen(ws, contracts: list[Contract], anomaly_list: list[dict]) -> None:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 6

    # --- Calculos en Python ---
    real = [c for c in contracts if not c.is_demo]
    demo = [c for c in contracts if c.is_demo]
    con_monto = [c for c in contracts if c.amount is not None]
    sin_monto = [c for c in contracts if c.amount is None]
    total_monto = sum((c.amount for c in con_monto), Decimal("0"))
    avg_monto = total_monto / Decimal(len(con_monto)) if con_monto else Decimal("0")

    anomaly_contract_ids = {a["contract_id"] for a in anomaly_list}

    by_institution: dict[str, tuple[int, Decimal]] = defaultdict(lambda: (0, Decimal("0")))
    for c in contracts:
        name = (c.institution.name if c.institution else "Desconocida")
        cnt, total = by_institution[name]
        by_institution[name] = (cnt + 1, total + (c.amount or Decimal("0")))

    by_supplier: dict[str, tuple[int, Decimal]] = defaultdict(lambda: (0, Decimal("0")))
    for c in contracts:
        name = (c.supplier.name if c.supplier else "Desconocido")
        cnt, total = by_supplier[name]
        by_supplier[name] = (cnt + 1, total + (c.amount or Decimal("0")))

    by_procedure: dict[str, int] = defaultdict(int)
    for c in contracts:
        key = c.procedure_type or "No especificado"
        by_procedure[key] += 1

    top_institutions = sorted(by_institution.items(), key=lambda x: x[1][1], reverse=True)[:10]
    top_suppliers = sorted(by_supplier.items(), key=lambda x: x[1][1], reverse=True)[:10]

    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")

    # --- Escritura ---
    row = 1

    def section_title(text: str, r: int) -> int:
        cell = ws.cell(row=r, column=1, value=text)
        cell.font = Font(name=FONT, size=11, bold=True, color=C_WHITE)
        cell.fill = _fill(C_DARK)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells(f"A{r}:C{r}")
        ws.row_dimensions[r].height = 22
        return r + 1

    def data_row(label: str, value, r: int, fmt: str = "", bold_val: bool = False) -> int:
        lbl = ws.cell(row=r, column=1, value=label)
        lbl.font = Font(name=FONT, size=9)
        lbl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        lbl.fill = _fill(C_ALT if r % 2 == 0 else C_WHITE)

        val = ws.cell(row=r, column=2, value=value)
        val.font = Font(name=FONT, size=9, bold=bold_val)
        val.alignment = Alignment(horizontal="right", vertical="center")
        val.fill = _fill(C_ALT if r % 2 == 0 else C_WHITE)
        if fmt:
            val.number_format = fmt

        for col in (1, 2):
            ws.cell(row=r, column=col).border = _border()
        ws.row_dimensions[r].height = 16
        return r + 1

    def blank(r: int) -> int:
        ws.row_dimensions[r].height = 8
        return r + 1

    # Encabezado del reporte
    title_cell = ws.cell(row=row, column=1,
        value="Reporte de Gasto Publico — Ecuador (Fuente: SERCOP OCDS)")
    title_cell.font = Font(name=FONT, size=13, bold=True, color=C_DARK)
    ws.merge_cells(f"A{row}:C{row}")
    ws.row_dimensions[row].height = 28
    row += 1

    gen_cell = ws.cell(row=row, column=1, value=f"Generado: {generated_at}")
    gen_cell.font = Font(name=FONT, size=8, color="6B7280")
    ws.merge_cells(f"A{row}:C{row}")
    ws.row_dimensions[row].height = 14
    row += 2

    # Bloque estadisticas generales
    row = section_title("Estadisticas generales", row)
    row = data_row("Total de contratos", len(contracts), row)
    row = data_row("  Contratos reales (SERCOP OCDS)", len(real), row)
    row = data_row("  Contratos demo / prueba", len(demo), row)
    row = data_row("Con monto registrado", len(con_monto), row)
    row = data_row("Sin monto (requieren enriquecimiento)", len(sin_monto), row)
    row = data_row("Monto total (USD)", float(total_monto), row, fmt='$#,##0.00', bold_val=True)
    row = data_row("Monto promedio (USD)", float(avg_monto), row, fmt='$#,##0.00')
    row = data_row("Contratos con alguna anomalia", len(anomaly_contract_ids), row, bold_val=True)
    row = blank(row)

    # Top 10 instituciones
    row = section_title("Top 10 instituciones por monto adjudicado (USD)", row)
    _sub_header(ws, row, ["Institucion", "Monto Total (USD)", "N° Contratos"])
    row += 1
    for inst_name, (cnt, total) in top_institutions:
        name_c = ws.cell(row=row, column=1, value=inst_name)
        name_c.font = Font(name=FONT, size=9)
        name_c.fill = _fill(C_ALT if row % 2 == 0 else C_WHITE)
        name_c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        name_c.border = _border()

        total_c = ws.cell(row=row, column=2, value=float(total))
        total_c.font = Font(name=FONT, size=9)
        total_c.fill = _fill(C_ALT if row % 2 == 0 else C_WHITE)
        total_c.number_format = '$#,##0.00'
        total_c.alignment = Alignment(horizontal="right", vertical="center")
        total_c.border = _border()

        cnt_c = ws.cell(row=row, column=3, value=cnt)
        cnt_c.font = Font(name=FONT, size=9)
        cnt_c.fill = _fill(C_ALT if row % 2 == 0 else C_WHITE)
        cnt_c.alignment = Alignment(horizontal="center", vertical="center")
        cnt_c.border = _border()

        ws.row_dimensions[row].height = 16
        row += 1
    row = blank(row)

    # Top 10 proveedores
    row = section_title("Top 10 proveedores por monto recibido (USD)", row)
    _sub_header(ws, row, ["Proveedor", "Monto Total (USD)", "N° Contratos"])
    row += 1
    for sup_name, (cnt, total) in top_suppliers:
        name_c = ws.cell(row=row, column=1, value=sup_name)
        name_c.font = Font(name=FONT, size=9)
        name_c.fill = _fill(C_ALT if row % 2 == 0 else C_WHITE)
        name_c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        name_c.border = _border()

        total_c = ws.cell(row=row, column=2, value=float(total))
        total_c.font = Font(name=FONT, size=9)
        total_c.fill = _fill(C_ALT if row % 2 == 0 else C_WHITE)
        total_c.number_format = '$#,##0.00'
        total_c.alignment = Alignment(horizontal="right", vertical="center")
        total_c.border = _border()

        cnt_c = ws.cell(row=row, column=3, value=cnt)
        cnt_c.font = Font(name=FONT, size=9)
        cnt_c.fill = _fill(C_ALT if row % 2 == 0 else C_WHITE)
        cnt_c.alignment = Alignment(horizontal="center", vertical="center")
        cnt_c.border = _border()

        ws.row_dimensions[row].height = 16
        row += 1
    row = blank(row)

    # Distribucion por tipo de procedimiento
    row = section_title("Distribucion por tipo de procedimiento", row)
    _sub_header(ws, row, ["Tipo de Procedimiento", "N° Contratos", ""])
    row += 1
    for proc_type, cnt in sorted(by_procedure.items(), key=lambda x: x[1], reverse=True):
        name_c = ws.cell(row=row, column=1, value=proc_type)
        name_c.font = Font(name=FONT, size=9)
        name_c.fill = _fill(C_ALT if row % 2 == 0 else C_WHITE)
        name_c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        name_c.border = _border()

        cnt_c = ws.cell(row=row, column=2, value=cnt)
        cnt_c.font = Font(name=FONT, size=9)
        cnt_c.fill = _fill(C_ALT if row % 2 == 0 else C_WHITE)
        cnt_c.alignment = Alignment(horizontal="center", vertical="center")
        cnt_c.border = _border()

        ws.row_dimensions[row].height = 16
        row += 1


def _sub_header(ws, row: int, labels: list[str]) -> None:
    for col_idx, label in enumerate(labels, start=1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.font = Font(name=FONT, size=9, bold=True, color=C_WHITE)
        cell.fill = _fill(C_GRAY_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border("334155")
    ws.row_dimensions[row].height = 18


# ------------------------------------------------------------------
# Hoja 3: Anomalias
# ------------------------------------------------------------------
HEADERS_ANOMALIAS = [
    ("#", 5),
    ("OCID / ID Externo", 36),
    ("Titulo del Contrato", 48),
    ("Institucion", 30),
    ("Proveedor", 30),
    ("Monto (USD)", 16),
    ("Tipo de Anomalia", 22),
    ("Severidad", 12),
    ("Descripcion", 52),
    ("Fuente Oficial (URL)", 20),
]


def _write_anomalias(
    ws,
    contracts: list[Contract],
    anomaly_reasons: dict[int, list[str]],
    anomaly_severity: dict[int, str],
) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    for col_idx, (label, width) in enumerate(HEADERS_ANOMALIAS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = _hdr(label, size=9)
        cell.fill = _fill(C_ANOMALY.replace("FEE2E2", "991B1B"))
        cell.alignment = _center()
        cell.border = _border("7F1D1D")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28

    idx = 1
    for contract in contracts:
        reasons = anomaly_reasons.get(contract.id, [])
        severity = anomaly_severity.get(contract.id, "medium")
        row_fill = _fill(C_ANOMALY) if severity == "high" else _fill(C_AMBER)

        ws.cell(row=idx + 1, column=1, value=idx).font = Font(name=FONT, size=9)
        ws.cell(row=idx + 1, column=2, value=contract.external_id).font = Font(name=FONT, size=9)
        ws.cell(row=idx + 1, column=3, value=contract.title).font = Font(name=FONT, size=9)
        ws.cell(row=idx + 1, column=4, value=contract.institution.name if contract.institution else "").font = Font(name=FONT, size=9)
        ws.cell(row=idx + 1, column=5, value=contract.supplier.name if contract.supplier else "").font = Font(name=FONT, size=9)

        amount_cell = ws.cell(row=idx + 1, column=6, value=float(contract.amount) if contract.amount is not None else None)
        amount_cell.font = Font(name=FONT, size=9)
        amount_cell.number_format = '#,##0.00'

        anomaly_type = "; ".join({a.split(" ")[0] for a in reasons})
        ws.cell(row=idx + 1, column=7, value=anomaly_type).font = Font(name=FONT, size=9)

        sev_cell = ws.cell(row=idx + 1, column=8, value=severity.upper())
        sev_cell.font = Font(name=FONT, size=9, bold=True,
                             color="B91C1C" if severity == "high" else "92400E")
        sev_cell.alignment = _center()

        ws.cell(row=idx + 1, column=9, value="; ".join(reasons)).font = Font(name=FONT, size=9)

        url_cell = ws.cell(row=idx + 1, column=10)
        if contract.source_url:
            url_cell.value = "Ver en SERCOP"
            url_cell.hyperlink = contract.source_url
            url_cell.font = Font(name=FONT, size=9, color=C_BLUE, underline="single")
            url_cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in range(1, 11):
            ws.cell(row=idx + 1, column=col).fill = row_fill
            ws.cell(row=idx + 1, column=col).border = _border()
        ws.row_dimensions[idx + 1].height = 16

        idx += 1

    if idx == 1:
        ws.cell(row=2, column=1,
                value="No se detectaron anomalias en los contratos actuales.").font = Font(name=FONT, size=9, color="6B7280")


# ------------------------------------------------------------------
# Main
# ------------------------------------------------------------------
def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Exporta contratos a XLSX trazable.")
    p.add_argument("--include-demo", action="store_true",
                   help="Incluir contratos de prueba/demo en el reporte")
    p.add_argument("--output", default=None,
                   help="Ruta de salida del archivo (por defecto: data/exports/contratos_TIMESTAMP.xlsx)")
    return p.parse_args()


def main() -> None:
    args = parse_args()
    init_db()

    with SessionLocal() as db:
        query = (
            db.query(Contract)
            .options(joinedload(Contract.institution), joinedload(Contract.supplier))
        )
        if not args.include_demo:
            query = query.filter(Contract.is_demo.is_(False))
        contracts = query.order_by(Contract.id.desc()).all()

    if not contracts:
        msg = "No hay contratos reales en la base. Usa --include-demo para incluir datos de prueba."
        print(msg)
        sys.exit(1)

    with SessionLocal() as db2:
        detect_contract_anomalies(db2)
        anomaly_list = _load_anomalies_from_db(db2)
    anomaly_ids: set[int] = {a["contract_id"] for a in anomaly_list}

    anomaly_reasons: dict[int, list[str]] = defaultdict(list)
    anomaly_severity: dict[int, str] = {}
    for a in anomaly_list:
        anomaly_reasons[a["contract_id"]].append(a["reason"])
        # severidad maxima gana
        prev = anomaly_severity.get(a["contract_id"], "medium")
        anomaly_severity[a["contract_id"]] = "high" if a["severity"] == "high" or prev == "high" else "medium"

    # Construir workbook
    wb = Workbook()

    ws_contratos = wb.active
    ws_contratos.title = "Contratos"
    _write_contratos(ws_contratos, contracts, anomaly_ids, anomaly_reasons, anomaly_severity)

    ws_resumen = wb.create_sheet("Resumen")
    _write_resumen(ws_resumen, contracts, anomaly_list)

    ws_anomalias = wb.create_sheet("Anomalias")
    anomalous = [c for c in contracts if c.id in anomaly_ids]
    _write_anomalias(ws_anomalias, anomalous, anomaly_reasons, anomaly_severity)

    # Guardar
    if args.output:
        output_path = Path(args.output)
    else:
        exports_dir = Path(__file__).resolve().parents[1] / "data" / "exports"
        exports_dir.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = exports_dir / f"contratos_sercop_{ts}.xlsx"

    wb.save(str(output_path))

    total_anomalias = len(anomaly_ids)
    print(f"EXPORTADO: {output_path}")
    print(f"Contratos incluidos : {len(contracts)}")
    print(f"Con anomalia        : {total_anomalias}")
    print(f"Pestaas generadas   : Contratos | Resumen | Anomalias")

    # Abrir automaticamente en Windows
    try:
        os.startfile(str(output_path))
    except (AttributeError, OSError):
        pass


if __name__ == "__main__":
    main()
